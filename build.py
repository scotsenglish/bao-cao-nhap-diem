#!/usr/bin/env python3
"""
Build script for Dashboard theo doi diem c-Learning - LMS.

Reads the monthly LMS export (data/latest.xlsx), computes all metrics,
and embeds the data into template.html to produce index.html.

Run: python build.py
Requires: pip install openpyxl
"""
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent
DATA_FILE = ROOT / "data" / "latest.xlsx"
TEMPLATE_FILE = ROOT / "template.html"
OUTPUT_FILE = ROOT / "index.html"

# ---------------------------------------------------------------------------
# Chi nhanh -> Vung mapping.
# Neu mo chi nhanh moi, them 1 dong vao day theo dung dinh dang, roi chay lai
# script nay (hoac de GitHub Actions tu chay).
# ---------------------------------------------------------------------------
REGION_MAP_RAW = """Vùng 1	Kim Giang
Vùng 1	Thanh Hóa
Vùng 1	Lam Sơn
Vùng 1	Nguyễn Xiển
Vùng 1	Linh Đàm
Vùng 1	Tây Hồ
Vùng 1	Nguyễn Tuân
Vùng 1	Hoàng Đạo Thúy
Vùng 1	Hoàng Quốc Việt
Vùng 1	Trung Văn
Vùng của Liên	Hải Dương
Vùng của Liên	Vĩnh Phúc
Vùng của Liên	Long Biên
Vùng của Liên	Vĩnh Phúc 3
Vùng của Liên	Phúc Yên
Vùng của Liên	Việt Trì
Vùng của Liên	Mỹ Đình
Vùng của Liên	Vinhomes Gardenia
Vùng 3	Times City
Vùng 3	Văn Khê
Vùng 3	An Khánh
Vùng 3	Vinhomes Smart City
Vùng 3	Vinhomes Smart City 2
Vùng 3	Dương Nội
Vùng 3	Phạm Văn Đồng
Vùng 3	Thái Bình
Vùng 5	Từ Sơn
Vùng 5	Hải Phòng 2
Vùng 5	Hải Phòng
Vùng 5	Bắc Ninh
Vùng 5	Bắc Ninh 2
Vùng 5	Bắc Giang
Vùng 6	Sài Đồng
Vùng 6	Vinh
Vùng 6	Ocean Park
Vùng 6	Trường Chinh
Vùng 6	Định Công
Vùng 7	Đà Nẵng
Vùng 7	Đà Nẵng 2
Vùng 7	Phan Văn Trị
Vùng 8	Celadon - Tân Phú
Vùng 8	Phạm Văn Chiêu
Vùng 8	Grand Park"""


def build_region_map():
    region_map = {}
    for line in REGION_MAP_RAW.strip().split("\n"):
        region, campus = line.split("\t")
        region_map["Scots English " + campus.strip()] = region.strip()
    return region_map


def num(v):
    return v if v is not None else 0


def extract(sheet_rows, idx, is_student, region_map):
    out = []
    for row in sheet_rows:
        branch = row[idx["Branch"]]
        if branch is None:
            continue
        hw_rec = num(row[idx["Homework Records"]])
        bt_rec = num(row[idx["Book Test Records"]])
        lq_rec = num(row[idx["Lesson Quiz Records"]])
        hw_in = num(row[idx["Homework Input Count"]])
        bt_in = num(row[idx["Book Test Input Count"]])
        lq_in = num(row[idx["Lesson Quiz Input Count"]])
        hw_st = row[idx["Homework Score Total"]]
        hw_sc = num(row[idx["Homework Score Count"]])
        bt_st = row[idx["Book Test Score Total"]]
        bt_sc = num(row[idx["Book Test Score Count"]])
        lq_st = row[idx["Lesson Quiz Score Total"]]
        lq_sc = num(row[idx["Lesson Quiz Score Count"]])

        rec = {
            "region": region_map.get(branch, "Chưa gán vùng"),
            "branch": branch,
            "program": row[idx["Program"]],
            "syllabus": row[idx["Syllabus"]],
            "class_name": row[idx["Class"]],
            "hw_records": hw_rec, "bt_records": bt_rec, "lq_records": lq_rec,
            "hw_input": hw_in, "bt_input": bt_in, "lq_input": lq_in,
            "hw_score_total": num(hw_st), "hw_score_count": hw_sc,
            "bt_score_total": num(bt_st), "bt_score_count": bt_sc,
            "lq_score_total": num(lq_st), "lq_score_count": lq_sc,
        }
        if is_student:
            name = row[idx["Name"]]
            staff = row[idx["Assigned Staff"]]
            rec["student_id"] = row[idx["ID"]]
            rec["name"] = name.strip() if name else ""
            rec["staff"] = staff.strip() if staff else "Chưa gán GV"
        else:
            rec["week"] = row[idx["Up To Week"]]
        out.append(rec)
    return out


def main():
    if not DATA_FILE.exists():
        print(f"LOI: Khong tim thay file du lieu tai {DATA_FILE}")
        print("Hay dam bao file Excel duoc dat dung duong dan: data/latest.xlsx")
        sys.exit(1)

    if not TEMPLATE_FILE.exists():
        print(f"LOI: Khong tim thay template.html tai {TEMPLATE_FILE}")
        sys.exit(1)

    region_map = build_region_map()

    wb = openpyxl.load_workbook(DATA_FILE, data_only=True)

    required_sheets = ["c-Learning Summary", "c-Learning Student Summary"]
    for s in required_sheets:
        if s not in wb.sheetnames:
            print(f"LOI: Khong tim thay sheet '{s}' trong file Excel.")
            print(f"Cac sheet hien co: {wb.sheetnames}")
            sys.exit(1)

    ws = wb["c-Learning Summary"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    class_rows = extract(list(ws.iter_rows(min_row=2, values_only=True)), idx, False, region_map)

    ws2 = wb["c-Learning Student Summary"]
    headers2 = [c.value for c in ws2[1]]
    idx2 = {h: i for i, h in enumerate(headers2)}
    student_rows = extract(list(ws2.iter_rows(min_row=2, values_only=True)), idx2, True, region_map)

    # Attach "week" (Up To Week) to each student via (branch, class_name) lookup
    week_lookup = {}
    for c in class_rows:
        week_lookup[(c["branch"], c["class_name"])] = c["week"]
    for s in student_rows:
        s["week"] = week_lookup.get((s["branch"], s["class_name"]))

    # Warn about branches missing from the region map
    unmapped = sorted({c["branch"] for c in class_rows if c["region"] == "Chưa gán vùng"})
    if unmapped:
        print("CANH BAO: Cac chi nhanh sau chua co trong REGION_MAP (se hien 'Chua gan vung'):")
        for b in unmapped:
            print(f"  - {b}")
        print("Them chi nhanh nay vao REGION_MAP_RAW trong build.py de gan dung Vung.\n")

    class_json = json.dumps(class_rows, ensure_ascii=False)
    student_json = json.dumps(student_rows, ensure_ascii=False)
    region_map_json = json.dumps(region_map, ensure_ascii=False)

    template = TEMPLATE_FILE.read_text(encoding="utf-8")
    output = (
        template
        .replace("__CLASS_DATA__", class_json)
        .replace("__STUDENT_DATA__", student_json)
        .replace("__REGION_MAP__", region_map_json)
    )

    OUTPUT_FILE.write_text(output, encoding="utf-8")
    print(f"OK: Da tao {OUTPUT_FILE} ({len(class_rows)} lop, {len(student_rows)} hoc vien)")


if __name__ == "__main__":
    main()
